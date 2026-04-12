"""
scripts/generate_fedex_batch.py

Generates FedEx batch claim Excel from queued FedEx claims.
Returns bytes (no file storage) for Streamlit download.
Max 200 claims per batch (FedEx limit).
"""

import io
import json
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.models import get_session, Claim, Order, Failure, FedExBatch

# ── FIXED VALUES FOR LATE CLAIMS ──────────────────────────────────
CLAIM_TYPE       = "Shipment not received"
ITEM_DESCRIPTION = "Fresh flower arrangement"
QTY              = 1
UNIT_COST        = 100.00
CURRENCY         = "U.S. Dollar (USD)"
COMMODITY        = "Food, Animal, Agricultural, and Medical products, both perishable and non-perishable"
MAX_PER_BATCH    = 200


def generate_batch_id() -> str:
    """Generate batch ID like FB-20260330-001."""
    session = get_session()
    try:
        today_str = datetime.now().strftime("%Y%m%d")
        prefix = f"FB-{today_str}-"
        existing = session.query(FedExBatch).filter(
            FedExBatch.batch_id.like(f"{prefix}%")
        ).count()
        seq = existing + 1
        return f"{prefix}{seq:03d}"
    finally:
        session.close()


def get_queued_fedex_count() -> int:
    """Count FedEx claims available for batching."""
    session = get_session()
    try:
        return session.query(Claim).filter(
            Claim.status == "queued_to_send",
            Claim.carrier == "FedEx",
        ).count()
    finally:
        session.close()


def create_batch(num_claims: int) -> dict:
    """
    Create a batch: assign batch_id to N FedEx claims.
    Returns {"batch_id": str, "claim_count": int, "claim_ids": list}
    """
    num_claims = min(num_claims, MAX_PER_BATCH)
    session = get_session()
    try:
        claims = session.query(Claim).filter(
            Claim.status == "queued_to_send",
            Claim.carrier == "FedEx",
        ).order_by(Claim.created_at).limit(num_claims).all()

        if not claims:
            return {"batch_id": None, "claim_count": 0, "claim_ids": []}

        batch_id = generate_batch_id()

        # Create batch record
        session.add(FedExBatch(
            batch_id=batch_id,
            claim_count=len(claims),
            status="ready",
        ))

        # Mark claims
        claim_ids = []
        for claim in claims:
            claim.status = "batch_downloaded"
            claim.fedex_batch_id = batch_id
            claim.updated_at = datetime.now()
            claim_ids.append(claim.claim_id)

        session.commit()
        return {"batch_id": batch_id, "claim_count": len(claims), "claim_ids": claim_ids}
    except Exception as e:
        session.rollback()
        return {"batch_id": None, "claim_count": 0, "error": str(e)}
    finally:
        session.close()


def generate_excel_bytes(batch_id: str) -> bytes:
    """
    Generate filled FedEx Excel for a batch. Returns bytes for download.
    Re-generates from DB every time — no file storage needed.
    """
    import openpyxl

    session = get_session()
    try:
        claims = session.query(Claim).filter(
            Claim.fedex_batch_id == batch_id
        ).order_by(Claim.created_at).all()

        if not claims:
            return b""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FedEx Batch Claims"

        # Row 1: Section headers
        ws.cell(row=1, column=1, value="Columns A-G are required. Column H is not required but recommended")
        ws.cell(row=1, column=9, value='Columns I, J, & L are required for "Missing Contents" or "Shipment Damaged"')
        ws.cell(row=1, column=14, value="Columns N-S are optional")

        # Row 2: Column headers
        headers = [
            "FedEx Tracking/PRO Number", "Claim Type", "Item Description",
            "Qty", "Unit Cost ($)", "Currency Type", "Commodity",
            "Shipping Cost ($)", "Type of Damage to Contents",
            "Type of Damage to Outer Packaging",
            "Other (Damage to Outer Packaging)", "Packaging Materials",
            "Other (Packaging Materials)", "Additional Comments",
            "Customer Reference Number", "Part", "Model",
            "Manufacturer", "Serial Number",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=2, column=i, value=h)

        # Row 3: Hints
        ws.cell(row=3, column=1, value="Max. 200 claims per spreadsheet")
        ws.cell(row=3, column=14, value="(Max 295 Characters)")

        # Data rows start at row 4
        for idx, claim in enumerate(claims):
            row = 4 + idx

            order = session.query(Order).filter_by(
                tracking_id=claim.tracking_id
            ).first()
            failure = session.query(Failure).filter_by(
                failure_id=claim.failure_id
            ).first()

            # A: Tracking Number
            ws.cell(row=row, column=1, value=str(claim.tracking_id))

            # B: Claim Type
            ws.cell(row=row, column=2, value=CLAIM_TYPE)

            # C: Item Description
            ws.cell(row=row, column=3, value=ITEM_DESCRIPTION)

            # D: Qty
            ws.cell(row=row, column=4, value=QTY)

            # E: Unit Cost
            ws.cell(row=row, column=5, value=UNIT_COST)

            # F: Currency
            ws.cell(row=row, column=6, value=CURRENCY)

            # G: Commodity
            ws.cell(row=row, column=7, value=COMMODITY)

            # H: Shipping Cost
            ws.cell(row=row, column=8, value=float(claim.claim_amount or 0) if claim.claim_amount else "")

            # N: Additional Comments (max 295 chars)
            comment = ""
            if claim.llm_narrative:
                comment = claim.llm_narrative[:295]
            elif failure:
                comment = f"Late delivery — {failure.delay_days or 1} day(s) past guaranteed date. Ship date: {failure.ship_date or ''}."
            ws.cell(row=row, column=14, value=comment)

            # O: Customer Reference
            ws.cell(row=row, column=15, value=str(order.partner_order_id if order else ""))

        # Write to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    finally:
        session.close()


def mark_batch_filed(batch_id: str, fedex_ref_id: str = "") -> dict:
    """Mark a batch as filed — all claims become filed_via_portal."""
    session = get_session()
    try:
        batch = session.query(FedExBatch).filter_by(batch_id=batch_id).first()
        if not batch:
            return {"success": False, "error": "Batch not found"}

        batch.status = "filed"
        batch.fedex_ref_id = fedex_ref_id or None

        claims = session.query(Claim).filter(
            Claim.fedex_batch_id == batch_id
        ).all()
        for claim in claims:
            claim.status = "filed_via_portal"
            claim.filed = True
            claim.filed_at = datetime.now()
            if fedex_ref_id:
                claim.carrier_case_id = fedex_ref_id
            claim.updated_at = datetime.now()

        session.commit()
        return {"success": True, "claims_filed": len(claims)}
    except Exception as e:
        session.rollback()
        return {"success": False, "error": str(e)}
    finally:
        session.close()


def discard_batch(batch_id: str) -> dict:
    """Discard a batch — claims go back to queued_to_send."""
    session = get_session()
    try:
        batch = session.query(FedExBatch).filter_by(batch_id=batch_id).first()
        if not batch:
            return {"success": False, "error": "Batch not found"}

        batch.status = "discarded"

        claims = session.query(Claim).filter(
            Claim.fedex_batch_id == batch_id
        ).all()
        for claim in claims:
            claim.status = "queued_to_send"
            claim.fedex_batch_id = None
            claim.updated_at = datetime.now()

        session.commit()
        return {"success": True, "claims_released": len(claims)}
    except Exception as e:
        session.rollback()
        return {"success": False, "error": str(e)}
    finally:
        session.close()


def get_all_batches() -> list:
    """Get all batches for display."""
    session = get_session()
    try:
        batches = session.query(FedExBatch).order_by(
            FedExBatch.created_at.desc()
        ).all()
        return [{
            "batch_id":     b.batch_id,
            "created_at":   b.created_at,
            "claim_count":  b.claim_count,
            "status":       b.status,
            "fedex_ref_id": b.fedex_ref_id or "",
            "notes":        b.notes or "",
        } for b in batches]
    finally:
        session.close()
